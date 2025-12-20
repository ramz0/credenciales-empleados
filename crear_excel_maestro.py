#!/usr/bin/env python3
"""
Script para crear Excel maestro combinando empleados actuales + nuevos gerentes.
The Money Center - Directorio de Empleados

Este script:
1. Lee empleados.json actual (con sus UUIDs)
2. Lee el Excel de gerentes nuevos
3. Genera un Excel maestro con TODOS los empleados
4. El Excel maestro incluye una columna UUID para preservar las URLs de QR
"""

import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import sys
import uuid
from pathlib import Path


def crear_excel_maestro(
    json_file='empleados.json',
    excel_gerentes=None,
    output_file='empleados_maestro.xlsx'
):
    """
    Crea un Excel maestro combinando empleados existentes + nuevos.

    Args:
        json_file: Archivo JSON con empleados actuales (con UUIDs)
        excel_gerentes: Archivo Excel con gerentes nuevos (opcional)
        output_file: Nombre del archivo Excel de salida
    """
    try:
        print("=" * 70)
        print("  CREADOR DE EXCEL MAESTRO - THE MONEY CENTER")
        print("=" * 70)
        print()

        # 1. Leer empleados actuales del JSON
        print(f"📂 Leyendo empleados actuales de: {json_file}")
        with open(json_file, 'r', encoding='utf-8') as f:
            empleados_actuales = json.load(f)

        print(f"   ✓ {len(empleados_actuales)} empleados cargados")

        # 2. Leer gerentes del Excel si se proporciona
        empleados_nuevos = []
        if excel_gerentes:
            print(f"\n📂 Leyendo gerentes nuevos de: {excel_gerentes}")
            wb = openpyxl.load_workbook(excel_gerentes)
            sheet = wb.active

            for idx, row in enumerate(sheet.iter_rows(min_row=4, values_only=True), start=4):
                # Columnas: B=gerencia (índice 1), C=nombre (índice 2), D=celular (índice 3)
                gerencia = row[1] if len(row) > 1 and row[1] else ""
                nombre = row[2] if len(row) > 2 and row[2] else ""
                celular = row[3] if len(row) > 3 and row[3] else ""

                # Saltar filas vacías
                if not nombre or not gerencia:
                    continue

                # Limpiar datos
                nombre = str(nombre).strip().upper()
                gerencia = str(gerencia).strip()
                celular = str(celular).strip()

                empleados_nuevos.append({
                    'id': str(uuid.uuid4()),  # Generar UUID nuevo
                    'nombre': nombre,
                    'puesto': 'Gerente',
                    'gerencia': gerencia,
                    'celular': celular
                })

            print(f"   ✓ {len(empleados_nuevos)} gerentes nuevos cargados")

        # 3. Crear Excel maestro
        print(f"\n📝 Creando Excel maestro...")

        # Crear nuevo workbook
        wb_maestro = openpyxl.Workbook()
        ws = wb_maestro.active
        ws.title = "Empleados"

        # Estilos para encabezado
        header_fill = PatternFill(start_color="EF4444", end_color="EF4444", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Encabezados
        headers = ["UUID", "NOMBRE", "PUESTO", "GERENCIA", "CELULAR"]
        ws.append(headers)

        # Aplicar estilos a encabezados
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Agregar empleados actuales
        for emp in empleados_actuales:
            ws.append([
                emp.get('id', ''),
                emp.get('nombre', ''),
                emp.get('puesto', ''),
                emp.get('gerencia', ''),
                emp.get('celular', '')
            ])

        # Agregar empleados nuevos
        for emp in empleados_nuevos:
            ws.append([
                emp.get('id', ''),
                emp.get('nombre', ''),
                emp.get('puesto', ''),
                emp.get('gerencia', ''),
                emp.get('celular', '')
            ])

        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 38  # UUID
        ws.column_dimensions['B'].width = 45  # NOMBRE
        ws.column_dimensions['C'].width = 15  # PUESTO
        ws.column_dimensions['D'].width = 25  # GERENCIA
        ws.column_dimensions['E'].width = 18  # CELULAR

        # Congelar primera fila (encabezados)
        ws.freeze_panes = 'A2'

        # Guardar archivo
        wb_maestro.save(output_file)

        # Resumen
        total_empleados = len(empleados_actuales) + len(empleados_nuevos)
        print(f"\n✅ Excel maestro creado exitosamente!")
        print("=" * 70)
        print(f"📊 RESUMEN:")
        print(f"   • Empleados actuales:  {len(empleados_actuales)}")
        print(f"   • Empleados nuevos:    {len(empleados_nuevos)}")
        print(f"   • TOTAL:               {total_empleados}")
        print(f"\n📁 Archivo generado: {output_file}")
        print("=" * 70)
        print()
        print("📋 PRÓXIMOS PASOS:")
        print("   1. Abre el Excel y verifica los datos")
        print("   2. Para agregar empleados: añade fila con UUID vacío")
        print("   3. Para actualizar: modifica datos pero mantén el UUID")
        print("   4. Para eliminar: borra la fila completa")
        print(f"   5. Ejecuta: python actualizar_empleados.py {output_file}")
        print()

        return True

    except FileNotFoundError as e:
        print(f"❌ Error: No se encontró el archivo: {e}")
        return False
    except Exception as e:
        print(f"❌ Error inesperado: {str(e)}")
        import traceback
        traceback.print_exc()
        return False


def main():
    """Función principal del script."""
    # Parámetros
    json_file = 'empleados.json'
    excel_gerentes = None
    output_file = 'empleados_maestro.xlsx'

    # Procesar argumentos
    if len(sys.argv) > 1:
        if sys.argv[1] in ['-h', '--help']:
            print("Uso: python crear_excel_maestro.py [excel_gerentes] [output_file]")
            print()
            print("Parámetros:")
            print("  excel_gerentes : Archivo Excel con gerentes nuevos (opcional)")
            print("  output_file    : Nombre del archivo de salida (default: empleados_maestro.xlsx)")
            print()
            print("Ejemplos:")
            print("  python crear_excel_maestro.py")
            print('  python crear_excel_maestro.py ~/Descargas/"GERENCIAS Y DR.xlsx"')
            print('  python crear_excel_maestro.py ~/Descargas/"GERENCIAS Y DR.xlsx" maestro.xlsx')
            sys.exit(0)

        excel_gerentes = sys.argv[1]

    if len(sys.argv) > 2:
        output_file = sys.argv[2]

    # Verificar que empleados.json existe
    if not Path(json_file).exists():
        print(f"❌ Error: El archivo '{json_file}' no existe")
        sys.exit(1)

    # Verificar que el Excel de gerentes existe (si se proporcionó)
    if excel_gerentes and not Path(excel_gerentes).exists():
        print(f"❌ Error: El archivo '{excel_gerentes}' no existe")
        sys.exit(1)

    # Ejecutar
    success = crear_excel_maestro(json_file, excel_gerentes, output_file)

    sys.exit(0 if success else 1)


if __name__ == '__main__':
    main()
