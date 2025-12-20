#!/usr/bin/env python3
"""
Script para limpiar y corregir el Excel maestro.
The Money Center - Directorio de Empleados

Este script:
1. Elimina filas con encabezados como registros
2. Corrige los puestos según la clasificación correcta
3. Ajusta gerencias para casos especiales
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import sys
from pathlib import Path


def limpiar_excel_maestro(input_file='empleados_maestro.xlsx', output_file='empleados_maestro_limpio.xlsx'):
    """
    Limpia y corrige el Excel maestro.

    Args:
        input_file: Excel maestro original
        output_file: Excel maestro corregido
    """
    try:
        print("=" * 70)
        print("  LIMPIEZA DE EXCEL MAESTRO - THE MONEY CENTER")
        print("=" * 70)
        print()
        print(f"📂 Leyendo: {input_file}")

        # Cargar Excel
        wb = openpyxl.load_workbook(input_file)
        ws = wb.active

        # Listas para tracking
        filas_eliminadas = []
        filas_corregidas = []

        # Crear nuevo workbook limpio
        wb_nuevo = openpyxl.Workbook()
        ws_nuevo = wb_nuevo.active
        ws_nuevo.title = "Empleados"

        # Estilos para encabezado
        header_fill = PatternFill(start_color="EF4444", end_color="EF4444", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Encabezados
        headers = ["UUID", "NOMBRE", "PUESTO", "GERENCIA", "CELULAR"]
        ws_nuevo.append(headers)

        # Aplicar estilos a encabezados
        for col_num, header in enumerate(headers, 1):
            cell = ws_nuevo.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Mapeo de nombres problemáticos para identificar encabezados
        nombres_invalidos = {
            'DIRECTOR', 'NOMBRE', 'NOMBRE DEL GERENTE',
            'GERENCIA', 'PUESTO', 'NO DE CELULAR'
        }

        # Mapeo de nombres a puestos correctos (casos especiales del Excel)
        directores = {
            'GERARDO MARÍN PÉREZ': ('Director', 'THE MONEY CENTER MSN'),
            'DAVID FRANCISCO GALICIA CAVAZOS': ('Director', 'THE MONEY CENTER MTY'),
            'ARTURO ELIZONDO': ('Director', 'THE MONEY CENTER MTY 1'),
        }

        asistentes_especiales = {
            'PATRICIA GONZALEZ MARTINEZ': ('Asistente de Préstamos', 'ADMINISTRACIÓN'),
            'THALIA EDITH BARRANCO ESPINOSA': ('Asistente de Dirección', 'DIRECCIÓN'),
        }

        consultores = {
            'EDUARDO GARCIA CHOMBO': ('Consultor', 'CONSULTORÍA'),
        }

        # Gerentes especiales (puesto en columna de gerencia)
        gerentes_especiales = {
            'KARLA PAOLA ALVARADO GALICIA': ('Gerente', 'THE MONEY CENTER MOD 40'),
        }

        print("\n🔍 Procesando empleados...")

        # Iterar filas (desde fila 2, saltando encabezado)
        fila_nueva = 2
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            uuid_val = row[0] if len(row) > 0 else ""
            nombre = row[1] if len(row) > 1 else ""
            puesto = row[2] if len(row) > 2 else ""
            gerencia = row[3] if len(row) > 3 else ""
            celular = row[4] if len(row) > 4 else ""

            # Saltar filas vacías
            if not nombre or str(nombre).strip() == "":
                continue

            nombre_limpio = str(nombre).strip().upper()

            # Eliminar filas con encabezados como nombres
            if nombre_limpio in nombres_invalidos:
                filas_eliminadas.append(f"Fila {idx}: {nombre_limpio}")
                print(f"  ❌ Eliminando encabezado: {nombre_limpio}")
                continue

            # Eliminar si gerencia es "DIRECCION" o "PUESTO" (encabezados)
            if str(gerencia).strip().upper() in ['DIRECCION', 'PUESTO', 'GERENCIA']:
                filas_eliminadas.append(f"Fila {idx}: {nombre_limpio} (gerencia={gerencia})")
                print(f"  ❌ Eliminando encabezado: {nombre_limpio}")
                continue

            # Corregir puestos según clasificación
            puesto_corregido = puesto
            gerencia_corregida = gerencia

            # Directores
            if nombre_limpio in directores:
                puesto_corregido, gerencia_corregida = directores[nombre_limpio]
                filas_corregidas.append(f"{nombre_limpio}: {puesto} → {puesto_corregido}")
                print(f"  ✏️  Corrigiendo: {nombre_limpio} → Director")

            # Asistentes
            elif nombre_limpio in asistentes_especiales:
                puesto_corregido, gerencia_corregida = asistentes_especiales[nombre_limpio]
                filas_corregidas.append(f"{nombre_limpio}: {puesto} → {puesto_corregido}")
                print(f"  ✏️  Corrigiendo: {nombre_limpio} → {puesto_corregido}")

            # Consultores
            elif nombre_limpio in consultores:
                puesto_corregido, gerencia_corregida = consultores[nombre_limpio]
                filas_corregidas.append(f"{nombre_limpio}: {puesto} → {puesto_corregido}")
                print(f"  ✏️  Corrigiendo: {nombre_limpio} → Consultor")

            # Gerentes especiales
            elif nombre_limpio in gerentes_especiales:
                puesto_corregido, gerencia_corregida = gerentes_especiales[nombre_limpio]
                filas_corregidas.append(f"{nombre_limpio}: gerencia → {gerencia_corregida}")
                print(f"  ✏️  Corrigiendo gerencia: {nombre_limpio}")

            # Gerentes de Money Centers (los que tienen gerencia tipo "THE MONEY CENTER X")
            elif puesto == "Gerente" and str(gerencia).startswith("THE MONEY CENTER"):
                puesto_corregido = "Gerente"
                # Mantener gerencia como está

            # Los demás son Asesores (mantener como están)
            else:
                puesto_corregido = puesto if puesto else "Asesor"

            # Agregar fila corregida
            ws_nuevo.append([
                str(uuid_val).strip() if uuid_val else "",
                nombre_limpio,
                puesto_corregido,
                str(gerencia_corregida).strip() if gerencia_corregida else "",
                str(celular).strip() if celular else ""
            ])

            fila_nueva += 1

        # Ajustar ancho de columnas
        ws_nuevo.column_dimensions['A'].width = 38  # UUID
        ws_nuevo.column_dimensions['B'].width = 45  # NOMBRE
        ws_nuevo.column_dimensions['C'].width = 25  # PUESTO
        ws_nuevo.column_dimensions['D'].width = 30  # GERENCIA
        ws_nuevo.column_dimensions['E'].width = 18  # CELULAR

        # Congelar primera fila
        ws_nuevo.freeze_panes = 'A2'

        # Guardar
        wb_nuevo.save(output_file)

        # Resumen
        total_final = fila_nueva - 2  # Restar encabezado y empezar desde 0

        print("\n✅ Limpieza completada!")
        print("=" * 70)
        print(f"📊 RESUMEN:")
        print(f"   • Filas eliminadas:    {len(filas_eliminadas)}")
        print(f"   • Filas corregidas:    {len(filas_corregidas)}")
        print(f"   • Total empleados:     {total_final}")
        print(f"\n📁 Archivo generado: {output_file}")
        print("=" * 70)

        if len(filas_eliminadas) > 0:
            print(f"\n🗑️  FILAS ELIMINADAS:")
            for fila in filas_eliminadas:
                print(f"   • {fila}")

        if len(filas_corregidas) > 0:
            print(f"\n✏️  CORRECCIONES REALIZADAS:")
            for correccion in filas_corregidas:
                print(f"   • {correccion}")

        print(f"\n📋 PRÓXIMOS PASOS:")
        print(f"   1. Revisa el archivo: {output_file}")
        print(f"   2. Si está correcto, reemplaza el original:")
        print(f"      mv {output_file} {input_file}")
        print(f"   3. Actualiza empleados.json:")
        print(f"      python actualizar_empleados.py {input_file}")
        print()

        return True

    except Exception as e:
        print(f"❌ Error: {str(e)}")
        import traceback
        traceback.print_exc()
        return False


def main():
    """Función principal."""
    input_file = 'empleados_maestro.xlsx'
    output_file = 'empleados_maestro_limpio.xlsx'

    if len(sys.argv) > 1:
        if sys.argv[1] in ['-h', '--help']:
            print("Uso: python limpiar_excel_maestro.py [input_file] [output_file]")
            print()
            print("Ejemplos:")
            print("  python limpiar_excel_maestro.py")
            print("  python limpiar_excel_maestro.py empleados_maestro.xlsx empleados_limpio.xlsx")
            sys.exit(0)
        input_file = sys.argv[1]

    if len(sys.argv) > 2:
        output_file = sys.argv[2]

    # Verificar que existe
    if not Path(input_file).exists():
        print(f"❌ Error: No se encontró '{input_file}'")
        sys.exit(1)

    # Ejecutar
    success = limpiar_excel_maestro(input_file, output_file)
    sys.exit(0 if success else 1)


if __name__ == '__main__':
    main()
